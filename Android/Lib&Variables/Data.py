from openpyxl import Workbook
from openpyxl import load_workbook
'''Author: Hoang'''
class Data:

    def get_data_by_sheet_name(self, sFilePath, sSheetName):
    #"""Get data from given file path, sheet name then return array of object""" 
        #from openpyxl import Workbook
        #from openpyxl import load_workbook
    
        wb = load_workbook(sFilePath.strip())
        ws = wb.get_sheet_by_name(sSheetName)
        row_count = ws.max_row
        column_count = ws.max_column
        #init data returned   
        test_data = []
    
        for i in range(2, row_count + 1):
            #ini dict
            row_data = {}
            for j in range(1, column_count + 1):
                if str(ws.cell(row = 1, column = j).value) == 'None':
                    continue
                else:
                    key = ws.cell(row = 1, column = j).value                 
                    value =  ws.cell(row = i, column = j).value
                    if value == None:
                        value = ''
                    row_data[key] = str(value)
            #print row_data
            test_data.append(row_data.copy())

        return test_data
    #end get_data_by_sheet_name 

    def get_row_count_by_sheet_name(self, sFilePath, sSheetName):
    #"""Count row data from given file path, sheet name then return array of object""" 
    
        wb = load_workbook(sFilePath.strip())
        ws = wb.get_sheet_by_name(sSheetName)
        row_count = ws.max_row
        #column_count = ws.max_column
        #init data returned   
        
        return row_count
    #end get_row_count_by_sheet_name 

    def get_column_count_by_sheet_name(self, sFilePath, sSheetName):
    #"""Count column data from given file path, sheet name then return array of object""" 
    
        wb = load_workbook(sFilePath.strip())
        ws = wb.get_sheet_by_name(sSheetName)
        #row_count = ws.max_row
        column_count = ws.max_column
        #init data returned   
        
        return column_count
    #end get_column_count_by_sheet_name 